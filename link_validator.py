import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import threading
import csv
import os
import time
from datetime import datetime
from utils import (
    detect_encoding, get_progress_interval, safe_str_conversion,
    validate_file_path, get_file_info, truncate_text
)
from config import current_config
from logger import get_logger
from exceptions import *


class LinkValidator:
    """Handles URL validation and file processing - simplified version"""
    
    def __init__(self):
        self.logger = get_logger()
        self.processing_cancelled = False
    
    def cancel_processing(self):
        """Cancel ongoing processing"""
        self.processing_cancelled = True
        self.logger.info("Processing cancellation requested")
    
    def is_valid_url(self, url):
        """Validate a URL using urllib.parse"""
        try:
            if pd.isna(url) or not str(url).strip():
                return False
                
            url_str = safe_str_conversion(url)
            if not url_str:
                return False
                
            result = urlparse(url_str)
            
            # Must have scheme and netloc
            if not result.scheme or not result.netloc:
                return False
                
            # Accept http/https schemes
            if result.scheme not in ('http', 'https'):
                return False
                
            # Basic netloc validation
            netloc = result.netloc.lower()
            if not netloc or netloc.startswith('.') or netloc.endswith('.'):
                return False
                
            return True
            
        except Exception as e:
            self.logger.debug(f"URL validation error for '{url}': {str(e)}")
            return False

    def process_csv(self, file_path, progress_callback=None):
        """Process CSV files - only process cells that look like URLs"""
        results = {'valid': [], 'invalid': []}
        start_time = time.time()
        
        try:
            validate_file_path(file_path)
            
            # Check file size
            file_info = get_file_info(file_path)
            if file_info.get('size_bytes', 0) > current_config.MAX_FILE_SIZE_MB * 1024 * 1024:
                raise FileSizeError(
                    current_config.ERROR_MESSAGES['file_too_large'],
                    file_path=file_path,
                    file_size=file_info.get('size_bytes', 0) / (1024 * 1024),
                    max_size=current_config.MAX_FILE_SIZE_MB
                )
            
            self.logger.log_file_processing_start(file_path, file_info)
            
            # Try to detect encoding
            encoding = detect_encoding(file_path)
            
            try:
                df = pd.read_csv(file_path, encoding=encoding)
            except UnicodeDecodeError:
                # Try fallback encodings
                for fallback_encoding in current_config.ENCODING_FALLBACKS:
                    try:
                        df = pd.read_csv(file_path, encoding=fallback_encoding, errors='ignore')
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    raise EncodingError(
                        current_config.ERROR_MESSAGES['encoding_failed'],
                        file_path=file_path,
                        tried_encodings=current_config.ENCODING_FALLBACKS
                    )
            
            total = len(df) * len(df.columns)
            progress_interval = get_progress_interval(total)
            processed = 0
            
            # Use sets to collect unique URLs
            valid_urls = set()
            invalid_urls = set()
            
            for column in df.columns:
                if self.processing_cancelled:
                    break
                    
                for index, value in df[column].items():
                    if pd.notna(value):
                        url_str = safe_str_conversion(value)
                        
                        # Only process cells that look like URLs
                        if url_str and self.looks_like_url(url_str):
                            if self.is_valid_url(url_str):
                                valid_urls.add(url_str)
                            else:
                                invalid_urls.add(url_str)
                    
                    processed += 1
                    if progress_callback and processed % progress_interval == 0:
                        progress_callback(processed/total)
            
            # Ensure final progress update shows 100%
            if progress_callback:
                progress_callback(1.0)
            
            # Convert sets to sorted lists
            results['valid'] = sorted(list(valid_urls))
            results['invalid'] = sorted(list(invalid_urls))
            
            processing_time = time.time() - start_time
            self.logger.info(f"CSV processing complete: {len(results['valid'])} valid, {len(results['invalid'])} invalid unique URLs")
            
        except Exception as e:
            if not isinstance(e, LinkValidatorError):
                self.logger.error(f"Unexpected CSV processing error: {str(e)}")
                raise FileProcessingError(f"CSV processing failed: {str(e)}", file_path=file_path, file_type='csv')
            raise
        
        return results

    def process_excel(self, file_path, progress_callback=None):
        """Process Excel files - only process cells that look like URLs"""
        results = {'valid': [], 'invalid': []}
        wb = None
        start_time = time.time()
        
        try:
            validate_file_path(file_path)
            
            file_info = get_file_info(file_path)
            if file_info.get('size_bytes', 0) > current_config.MAX_FILE_SIZE_MB * 1024 * 1024:
                raise FileSizeError(
                    current_config.ERROR_MESSAGES['file_too_large'],
                    file_path=file_path,
                    file_size=file_info.get('size_bytes', 0) / (1024 * 1024),
                    max_size=current_config.MAX_FILE_SIZE_MB
                )
            
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            
            # Calculate total cells
            total_cells = 0
            for ws in wb.worksheets:
                if ws.max_row and ws.max_column:
                    actual_rows = min(ws.max_row, current_config.MAX_EXCEL_ROWS)
                    actual_cols = min(ws.max_column, current_config.MAX_EXCEL_COLS)
                    total_cells += actual_rows * actual_cols
            
            progress_interval = get_progress_interval(total_cells)
            processed = 0
            
            # Use sets to collect unique URLs
            valid_urls = set()
            invalid_urls = set()
            
            for sheet in wb.worksheets:
                if self.processing_cancelled:
                    break
                    
                for row_num, row in enumerate(sheet.iter_rows(values_only=True, max_row=current_config.MAX_EXCEL_ROWS), 1):
                    if self.processing_cancelled:
                        break
                        
                    for col_num, cell_value in enumerate(row[:current_config.MAX_EXCEL_COLS], 1):
                        if cell_value:
                            cell_str = safe_str_conversion(cell_value)
                            
                            # Only process cells that look like URLs
                            if cell_str and self.looks_like_url(cell_str):
                                if self.is_valid_url(cell_str):
                                    valid_urls.add(cell_str)
                                else:
                                    invalid_urls.add(cell_str)
                        
                        processed += 1
                        if progress_callback and processed % progress_interval == 0:
                            progress_callback(min(processed/total_cells, 1.0))
            
            # Ensure final progress update shows 100%
            if progress_callback:
                progress_callback(1.0)
            
            # Convert sets to sorted lists
            results['valid'] = sorted(list(valid_urls))
            results['invalid'] = sorted(list(invalid_urls))
            
            processing_time = time.time() - start_time
            self.logger.info(f"Excel processing complete: {len(results['valid'])} valid, {len(results['invalid'])} invalid unique URLs")
            
        except Exception as e:
            if not isinstance(e, LinkValidatorError):
                self.logger.error(f"Unexpected Excel processing error: {str(e)}")
                raise FileProcessingError(f"Excel processing failed: {str(e)}", file_path=file_path, file_type='excel')
            raise
        finally:
            if wb:
                wb.close()
        
        return results

    def looks_like_url(self, text):
        """Check if text looks like a URL before validation - improved logic"""
        text = text.strip()
        
        # Empty or very short text
        if len(text) < 4:
            return False
        
        # Should not be too long (likely paragraph text)
        if len(text) > 200:
            return False
        
        # Should not contain too many spaces (likely sentence)
        if text.count(' ') > 2:
            return False
        
        # Check for obvious URL patterns first
        url_indicators = ['http://', 'https://', 'www.', 'ftp://']
        if any(indicator in text.lower() for indicator in url_indicators):
            return True
        
        # Check for domain-like patterns (contains dots and looks like domain)
        if '.' in text:
            # Split by spaces to handle cases where URL might be in a sentence
            words = text.split()
            for word in words:
                word = word.strip('.,!?()[]{}"\'-')  # Remove common punctuation
                
                # Check if word looks like a domain or URL
                if self.looks_like_domain_or_url(word):
                    return True
        
        # Check for common sentence indicators that would disqualify it
        sentence_indicators = [
            '. ', '! ', '? ', ', and ', ', or ', ' the ', ' a ', ' an ', 
            ' is ', ' are ', ' was ', ' were ', ' will ', ' can ', ' could ',
            ' should ', ' would ', ' may ', ' might ', ' this ', ' that ',
            ' these ', ' those ', ' with ', ' without ', ' from ', ' into ',
            ' about ', ' through ', ' during ', ' before ', ' after ', ' over ',
            ' under ', ' above ', ' below ', ' between ', ' among ', ' within '
        ]
        
        text_lower = text.lower()
        if any(indicator in text_lower for indicator in sentence_indicators):
            return False
        
        return False
    
    def looks_like_url(self, text):
        """Check if text looks like a URL before validation - comprehensive logic"""
        text = text.strip()
        
        # Empty or very short text
        if len(text) < 4:
            return False
        
        # Should not be too long (likely paragraph text)
        if len(text) > 200:
            return False
        
        # Should not contain too many spaces (likely sentence)
        if text.count(' ') > 2:
            return False
        
        # Check for obvious URL patterns first
        url_indicators = ['http://', 'https://', 'www.', 'ftp://']
        if any(indicator in text.lower() for indicator in url_indicators):
            return True
        
        # Check for common sentence indicators that would disqualify it
        sentence_indicators = [
            '. ', '! ', '? ', ', and ', ', or ', ' the ', ' a ', ' an ', 
            ' is ', ' are ', ' was ', ' were ', ' will ', ' can ', ' could ',
            ' should ', ' would ', ' may ', ' might ', ' this ', ' that ',
            ' these ', ' those ', ' with ', ' without ', ' from ', ' into ',
            ' about ', ' through ', ' during ', ' before ', ' after ', ' over ',
            ' under ', ' above ', ' below ', ' between ', ' among ', ' within '
        ]
        
        text_lower = text.lower()
        if any(indicator in text_lower for indicator in sentence_indicators):
            return False
        
        # Check for domain-like patterns (contains dots and looks like domain)
        if '.' in text:
            # Split by spaces to handle cases where URL might be in a sentence
            words = text.split()
            for word in words:
                word = word.strip('.,!?()[]{}"\'-')  # Remove common punctuation
                
                # Check if word looks like a domain or URL
                if self.looks_like_domain_or_url(word):
                    return True
        
        return False
    
    def looks_like_domain_or_url(self, word):
        """Check if a word looks like a domain name or URL - more permissive"""
        if not word or len(word) < 4:
            return False
        
        # Must contain at least one dot
        if '.' not in word:
            return False
        
        # Handle URLs with paths (like arxiv.org/ai-safety)
        base_part = word
        if '/' in word:
            base_part = word.split('/')[0]  # Get just the domain part
        
        # Split by dots
        parts = base_part.split('.')
        
        # Must have at least 2 parts for domain
        if len(parts) < 2:
            return False
        
        # Check if it looks like a domain structure
        # Allow more flexibility in TLD length (2-20 characters to handle things like "missing-domain")
        tld = parts[-1]
        if not (2 <= len(tld) <= 20):
            return False
        
        # TLD should contain mostly letters (allow hyphens for compound TLDs)
        if not all(c.isalpha() or c == '-' for c in tld):
            return False
        
        # All parts before TLD should contain only valid domain characters
        for part in parts[:-1]:  # All parts except the last (TLD)
            if not part:  # Empty part (like in "example..com")
                return False
            
            # Allow letters, numbers, and hyphens (but not at start/end)
            for char in part:
                if not (char.isalnum() or char == '-'):
                    return False
            
            # Hyphens cannot be at start or end
            if part.startswith('-') or part.endswith('-'):
                return False
        
        # Additional checks
        # Domain should have reasonable length
        if len(base_part) > 100:
            return False
        
        # Should not start or end with dot
        if base_part.startswith('.') or base_part.endswith('.'):
            return False
        
        # Must have at least one alphabetic character in the main domain part
        main_domain = parts[-2] if len(parts) >= 2 else parts[0]
        if not any(c.isalpha() for c in main_domain):
            return False
        
        return True

    def process_text(self, file_path, progress_callback=None):
        """Process text files - only process lines that look like URLs"""
        results = {'valid': [], 'invalid': []}
        start_time = time.time()
        
        try:
            validate_file_path(file_path)
            
            file_info = get_file_info(file_path)
            if file_info.get('size_bytes', 0) > current_config.MAX_FILE_SIZE_MB * 1024 * 1024:
                raise FileSizeError(
                    current_config.ERROR_MESSAGES['file_too_large'],
                    file_path=file_path,
                    file_size=file_info.get('size_bytes', 0) / (1024 * 1024),
                    max_size=current_config.MAX_FILE_SIZE_MB
                )
            
            # Detect encoding
            encoding = detect_encoding(file_path)
            encodings = [encoding] + current_config.ENCODING_FALLBACKS
            lines = None
            
            for enc in encodings:
                try:
                    with open(file_path, 'r', encoding=enc) as f:
                        lines = f.readlines()
                    break
                except UnicodeDecodeError:
                    continue
            
            if lines is None:
                raise EncodingError(
                    current_config.ERROR_MESSAGES['encoding_failed'],
                    file_path=file_path,
                    tried_encodings=encodings
                )
            
            total = len(lines)
            progress_interval = get_progress_interval(total)
            
            # Use sets to collect unique URLs
            valid_urls = set()
            invalid_urls = set()
            
            for i, line in enumerate(lines):
                if self.processing_cancelled:
                    break
                    
                line_str = line.strip()
                
                # Only process lines that look like URLs
                if line_str and self.looks_like_url(line_str):
                    if self.is_valid_url(line_str):
                        valid_urls.add(line_str)
                    else:
                        invalid_urls.add(line_str)
                
                if progress_callback and i % progress_interval == 0:
                    progress_callback(i/total)
            
            # Ensure final progress update shows 100%
            if progress_callback:
                progress_callback(1.0)
            
            # Ensure final progress update shows 100%
            if progress_callback:
                progress_callback(1.0)
            
            # Convert sets to sorted lists
            results['valid'] = sorted(list(valid_urls))
            results['invalid'] = sorted(list(invalid_urls))
            
            processing_time = time.time() - start_time
            self.logger.info(f"Text processing complete: {len(results['valid'])} valid, {len(results['invalid'])} invalid unique URLs")
            
        except Exception as e:
            if not isinstance(e, LinkValidatorError):
                self.logger.error(f"Unexpected text processing error: {str(e)}")
                raise FileProcessingError(f"Text file processing failed: {str(e)}", file_path=file_path, file_type='text')
            raise
        
        return results

    def process_html(self, file_path, progress_callback=None):
        """Process HTML files - simplified to collect unique URLs only"""
        results = {'valid': [], 'invalid': []}
        soup = None
        start_time = time.time()
        
        try:
            validate_file_path(file_path)
            
            file_info = get_file_info(file_path)
            if file_info.get('size_bytes', 0) > current_config.MAX_FILE_SIZE_MB * 1024 * 1024:
                raise FileSizeError(
                    current_config.ERROR_MESSAGES['file_too_large'],
                    file_path=file_path,
                    file_size=file_info.get('size_bytes', 0) / (1024 * 1024),
                    max_size=current_config.MAX_FILE_SIZE_MB
                )
            
            # Detect encoding
            encoding = detect_encoding(file_path)
            
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
            
            soup = BeautifulSoup(content, 'html.parser')
            elements = soup.find_all(['a', 'img', 'link', 'script'])
            total = len(elements)
            progress_interval = get_progress_interval(total)
            
            # Use sets to collect unique URLs
            valid_urls = set()
            invalid_urls = set()
            
            for i, element in enumerate(elements):
                if self.processing_cancelled:
                    break
                    
                url = element.get('href') or element.get('src')
                if url:
                    url_str = safe_str_conversion(url)
                    if url_str:
                        if self.is_valid_url(url_str):
                            valid_urls.add(url_str)
                        else:
                            invalid_urls.add(url_str)
                
                if progress_callback and i % progress_interval == 0:
                    progress_callback(i/total)
            
            # Convert sets to sorted lists
            results['valid'] = sorted(list(valid_urls))
            results['invalid'] = sorted(list(invalid_urls))
            
            processing_time = time.time() - start_time
            self.logger.info(f"HTML processing complete: {len(results['valid'])} valid, {len(results['invalid'])} invalid unique URLs")
            
        except Exception as e:
            if not isinstance(e, LinkValidatorError):
                self.logger.error(f"Unexpected HTML processing error: {str(e)}")
                raise FileProcessingError(f"HTML processing failed: {str(e)}", file_path=file_path, file_type='html')
            raise
        finally:
            if soup:
                soup.decompose()
        
        return results


class LinkValidatorApp:
    """Simplified GUI application - just lists of URLs"""
    
    def __init__(self, root):
        self.root = root
        self.root.title(f"{current_config.APP_NAME} v{current_config.APP_VERSION} - Simplified")
        self.root.geometry(f"{current_config.WINDOW_WIDTH}x{current_config.WINDOW_HEIGHT}")
        self.root.configure(bg=current_config.COLORS['background'])
        
        self.results = None
        self.current_file_info = None
        self.validator = LinkValidator()
        self.processing_thread = None
        self.logger = get_logger()
        
        self.create_widgets()
        self.center_window()
        
        self.logger.info(f"Link Validator v{current_config.APP_VERSION} (Simplified) started")

    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def create_widgets(self):
        """Create simplified GUI"""
        # Main frame
        main_frame = tk.Frame(self.root, padx=20, pady=20, bg=current_config.COLORS['background'])
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_frame = tk.Frame(main_frame, bg=current_config.COLORS['background'])
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(
            header_frame, 
            text=f"{current_config.APP_NAME} v{current_config.APP_VERSION}", 
            font=(current_config.FONT_FAMILY, 20, 'bold'),
            bg=current_config.COLORS['background'],
            fg=current_config.COLORS['primary']
        )
        title_label.pack(side=tk.LEFT)
        
        subtitle_label = tk.Label(
            header_frame, 
            text="Simplified - URL Lists Only", 
            font=(current_config.FONT_FAMILY, 11, 'italic'),
            bg=current_config.COLORS['background'],
            fg=current_config.COLORS['info']
        )
        subtitle_label.pack(side=tk.RIGHT)
        
        # File info
        file_info_frame = tk.Frame(main_frame, bg=current_config.COLORS['background'])
        file_info_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(
            file_info_frame, 
            text="📁 File:", 
            font=(current_config.FONT_FAMILY, 11, 'bold'),
            bg=current_config.COLORS['background']
        ).pack(side=tk.LEFT)
        
        self.file_info_label = tk.Label(
            file_info_frame, 
            text="No file selected", 
            font=(current_config.FONT_FAMILY, 11),
            bg=current_config.COLORS['background'],
            fg=current_config.COLORS['text']
        )
        self.file_info_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # Buttons
        button_frame = tk.Frame(main_frame, bg=current_config.COLORS['background'])
        button_frame.pack(fill=tk.X, pady=10)
        
        self.upload_btn = tk.Button(
            button_frame, 
            text="📂 Upload File", 
            command=self.start_file_processing,
            width=18,
            font=(current_config.FONT_FAMILY, 11, 'bold'),
            bg=current_config.COLORS['primary'],
            fg='white',
            relief=tk.FLAT
        )
        self.upload_btn.pack(side=tk.LEFT, padx=5)
        
        self.export_btn = tk.Button(
            button_frame, 
            text="💾 Export Results", 
            command=self.export_results,
            state=tk.DISABLED,
            width=18,
            font=(current_config.FONT_FAMILY, 11),
            bg=current_config.COLORS['success'],
            fg='white',
            relief=tk.FLAT
        )
        self.export_btn.pack(side=tk.LEFT, padx=5)
        
        self.clear_btn = tk.Button(
            button_frame,
            text="🗑️ Clear Results",
            command=self.clear_results,
            width=18,
            font=(current_config.FONT_FAMILY, 11),
            bg=current_config.COLORS['warning'],
            fg='white',
            relief=tk.FLAT
        )
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        
        # Cancel button (hidden initially)
        self.cancel_btn = tk.Button(
            button_frame,
            text="❌ Cancel",
            command=self.cancel_processing,
            width=18,
            font=(current_config.FONT_FAMILY, 11),
            bg=current_config.COLORS['error'],
            fg='white',
            relief=tk.FLAT
        )
        
        # Progress
        progress_frame = tk.Frame(main_frame, bg=current_config.COLORS['background'])
        progress_frame.pack(fill=tk.X, pady=15)
        
        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress.pack(fill=tk.X)
        
        progress_info_frame = tk.Frame(progress_frame, bg=current_config.COLORS['background'])
        progress_info_frame.pack(fill=tk.X, pady=(8, 0))
        
        self.progress_label = tk.Label(
            progress_info_frame, 
            text="Ready to process files", 
            font=(current_config.FONT_FAMILY, 10),
            bg=current_config.COLORS['background']
        )
        self.progress_label.pack(side=tk.LEFT)
        
        self.progress_stats = tk.Label(
            progress_info_frame,
            text="",
            font=(current_config.FONT_FAMILY, 10),
            bg=current_config.COLORS['background'],
            fg=current_config.COLORS['info']
        )
        self.progress_stats.pack(side=tk.RIGHT)
        
        # Results display
        results_frame = tk.Frame(main_frame, bg=current_config.COLORS['background'])
        results_frame.pack(fill=tk.BOTH, expand=True, pady=(15, 0))
        
        tk.Label(
            results_frame,
            text="📊 Results:",
            font=(current_config.FONT_FAMILY, 12, 'bold'),
            bg=current_config.COLORS['background']
        ).pack(anchor=tk.W, pady=(0, 10))
        
        self.output_text = scrolledtext.ScrolledText(
            results_frame, 
            wrap=tk.WORD, 
            width=110, 
            height=30,
            font=(current_config.MONOSPACE_FONT, 10),
            bg='#ffffff',
            fg=current_config.COLORS['text'],
            selectbackground=current_config.COLORS['primary'],
            selectforeground='white',
            relief=tk.FLAT,
            borderwidth=2
        )
        self.output_text.pack(fill=tk.BOTH, expand=True)
        
        # Status bar
        status_frame = tk.Frame(main_frame, bg=current_config.COLORS['background'])
        status_frame.pack(fill=tk.X, pady=(15, 0))
        
        self.status_bar = tk.Label(
            status_frame, 
            text="🟢 Ready - Select a file to begin URL validation", 
            bd=1, 
            relief=tk.SUNKEN, 
            anchor=tk.W,
            font=(current_config.FONT_FAMILY, 10),
            bg='white',
            padx=10,
            pady=3
        )
        self.status_bar.pack(fill=tk.X)

    def start_file_processing(self):
        """Start file processing with simplified dialog"""
        try:
            # Simple file dialog
            file_path = filedialog.askopenfilename(
                title="Select file to validate URLs",
                filetypes=[
                    ("All Files", "*.*"),
                    ("Text Files", "*.txt"),
                    ("CSV Files", "*.csv"),
                    ("Excel Files", "*.xlsx"),
                    ("HTML Files", "*.html")
                ]
            )
            
            if not file_path:
                return
            
            # Get file info
            validate_file_path(file_path)
            self.current_file_info = get_file_info(file_path)
            
            if 'error' in self.current_file_info:
                raise FileValidationError(f"Cannot read file: {self.current_file_info['error']}", file_path=file_path)
            
            # Update display
            info_text = f"{self.current_file_info['name']} ({self.current_file_info['size_formatted']})"
            if self.current_file_info.get('is_large'):
                info_text += " • ⚠️ Large File"
            
            self.file_info_label.config(text=info_text)
            self.clear_results()
            
            # Show processing info
            self.output_text.insert(tk.END, f"🔍 PROCESSING: {self.current_file_info['name']}\n")
            self.output_text.insert(tk.END, f"📏 Size: {self.current_file_info['size_formatted']}\n")
            self.output_text.insert(tk.END, f"⏰ Started: {datetime.now().strftime('%H:%M:%S')}\n\n")
            self.output_text.see(tk.END)
            
            self.update_status("🔄 Processing file...")
            self.set_processing_state(True)
            
            # Start processing thread
            self.processing_thread = threading.Thread(target=self.process_file, args=(file_path,), daemon=True)
            self.processing_thread.start()
            
        except Exception as e:
            self.logger.error(f"Error in file selection: {str(e)}")
            messagebox.showerror("Error", f"Failed to process file: {str(e)}")
            self.update_status("❌ Error occurred")

    def set_processing_state(self, processing):
        """Update UI state"""
        if processing:
            self.upload_btn.config(state=tk.DISABLED)
            self.export_btn.config(state=tk.DISABLED)
            self.clear_btn.config(state=tk.DISABLED)
            self.cancel_btn.pack(side=tk.LEFT, padx=5)
        else:
            self.upload_btn.config(state=tk.NORMAL)
            self.clear_btn.config(state=tk.NORMAL)
            self.cancel_btn.pack_forget()
            if self.results:
                self.export_btn.config(state=tk.NORMAL)

    def cancel_processing(self):
        """Cancel processing"""
        if self.validator:
            self.validator.cancel_processing()
            self.update_status("🛑 Cancelling...")
            self.progress_label.config(text="Cancelling...")

    def process_file(self, file_path):
        """Process file in background"""
        start_time = time.time()
        
        try:
            self.validator.processing_cancelled = False
            
            # Determine processor
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.csv':
                processor = self.validator.process_csv
            elif file_ext in ('.xls', '.xlsx'):
                processor = self.validator.process_excel
            elif file_ext in ('.html', '.htm', '.xml'):
                processor = self.validator.process_html
            else:
                processor = self.validator.process_text
            
            # Process
            results = processor(file_path, self.update_progress)
            
            if self.validator.processing_cancelled:
                self.root.after(0, lambda: self.output_text.insert(tk.END, "\n❌ PROCESSING CANCELLED\n"))
                self.update_status("❌ Cancelled")
                return
            
            # Ensure progress bar shows 100% when complete
            self.root.after(0, lambda: self.progress.config(value=100))
            self.root.after(0, lambda: self.progress_label.config(text="Complete - 100%"))
            
            processing_time = time.time() - start_time
            self.results = results
            self.root.after(0, lambda: self.display_results(results, processing_time))
            self.update_status(f"✅ Completed in {processing_time:.1f}s")
            
        except Exception as e:
            error_msg = f"Processing error: {str(e)}"
            self.root.after(0, lambda: self.output_text.insert(tk.END, f"\n❌ ERROR: {error_msg}\n"))
            self.update_status(f"❌ Error: {str(e)}")
            # Ensure progress bar shows completion even on error
            self.root.after(0, lambda: self.progress.config(value=100))
            self.root.after(0, lambda: self.progress_label.config(text="Error"))
        finally:
            self.root.after(0, lambda: self.set_processing_state(False))

    def update_progress(self, progress):
        """Update progress bar"""
        try:
            percent = max(0, min(100, int(progress * 100)))
            
            if hasattr(self, '_progress_start_time'):
                elapsed = time.time() - self._progress_start_time
                if progress > 0:
                    estimated_total = elapsed / progress
                    remaining = estimated_total - elapsed
                    remaining_str = f" • {remaining:.0f}s remaining" if remaining > 0 else ""
                else:
                    remaining_str = ""
            else:
                self._progress_start_time = time.time()
                remaining_str = ""
            
            self.root.after(0, lambda: self.progress.config(value=percent))
            self.root.after(0, lambda: self.progress_label.config(text=f"Processing... {percent}%"))
            self.root.after(0, lambda: self.progress_stats.config(text=remaining_str))
            
        except Exception as e:
            self.logger.debug(f"Progress update error: {e}")

    def display_results(self, results, processing_time):
        """Display simplified results - just lists of URLs"""
        valid_count = len(results['valid'])
        invalid_count = len(results['invalid'])
        total_count = valid_count + invalid_count
        
        # Header
        self.output_text.insert(tk.END, "🏁 PROCESSING COMPLETE\n")
        self.output_text.insert(tk.END, "="*80 + "\n\n")
        
        # Summary
        self.output_text.insert(tk.END, "📊 SUMMARY\n")
        self.output_text.insert(tk.END, f"   ⏱️  Processing time: {processing_time:.2f} seconds\n")
        self.output_text.insert(tk.END, f"   📈 Total unique URLs: {total_count:,}\n")
        self.output_text.insert(tk.END, f"   ✅ Valid URLs: {valid_count:,}\n")
        self.output_text.insert(tk.END, f"   ❌ Invalid URLs: {invalid_count:,}\n")
        
        if total_count > 0:
            success_rate = (valid_count / total_count) * 100
            self.output_text.insert(tk.END, f"   📊 Success rate: {success_rate:.1f}%\n")
        
        self.output_text.insert(tk.END, "\n")
        
        # Valid URLs
        if results['valid']:
            self.output_text.insert(tk.END, "✅ VALID URLS\n")
            self.output_text.insert(tk.END, "-" * 50 + "\n")
            for i, url in enumerate(results['valid'], 1):
                self.output_text.insert(tk.END, f"{i:3d}. {url}\n")
            self.output_text.insert(tk.END, "\n")
        
        # Invalid URLs
        if results['invalid']:
            self.output_text.insert(tk.END, "❌ INVALID URLS\n")
            self.output_text.insert(tk.END, "-" * 50 + "\n")
            for i, url in enumerate(results['invalid'], 1):
                self.output_text.insert(tk.END, f"{i:3d}. {url}\n")
            self.output_text.insert(tk.END, "\n")
        
        if not results['valid'] and not results['invalid']:
            self.output_text.insert(tk.END, "🤷 No URLs found in the file.\n")
        elif not results['invalid']:
            self.output_text.insert(tk.END, "🎉 All URLs are valid!\n")
        
        self.output_text.insert(tk.END, "="*80 + "\n")
        self.output_text.insert(tk.END, f"✨ Completed at {datetime.now().strftime('%H:%M:%S')}\n\n")
        self.output_text.see(tk.END)

    def clear_results(self):
        """Clear all results"""
        self.output_text.delete(1.0, tk.END)
        self.results = None
        self.export_btn.config(state=tk.DISABLED)
        self.progress['value'] = 0
        self.progress_label.config(text="Ready to process files")
        self.progress_stats.config(text="")
        if hasattr(self, '_progress_start_time'):
            delattr(self, '_progress_start_time')

    def export_results(self):
        """Export simplified results"""
        if not self.results:
            messagebox.showwarning("No Results", "No results to export")
            return
        
        if not self.results['valid'] and not self.results['invalid']:
            messagebox.showwarning("Empty Results", "No URLs found to export")
            return
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = self.current_file_info['name'] if self.current_file_info else "unknown"
        file_base = os.path.splitext(file_name)[0]
        default_filename = f"url_validation_{file_base}_{timestamp}.csv"
        
        save_path = filedialog.asksaveasfilename(
            title="Save URL validation results",
            defaultextension=".csv",
            filetypes=[
                ("CSV Files", "*.csv"),
                ("Text Files", "*.txt"),
                ("All Files", "*.*")
            ],
            initialfile=default_filename
        )
        
        if not save_path:
            return
        
        try:
            export_format = os.path.splitext(save_path)[1].lower()
            
            if export_format == '.csv' or export_format == '':
                self._export_csv(save_path if save_path.endswith('.csv') else save_path + '.csv')
            elif export_format == '.txt':
                self._export_text(save_path)
            else:
                self._export_csv(save_path)
            
            # Get actual file path
            actual_path = save_path if os.path.exists(save_path) else save_path + '.csv'
            total_urls = len(self.results['valid']) + len(self.results['invalid'])
            file_size = os.path.getsize(actual_path)
            
            success_msg = (f"✅ Export completed successfully!\n\n"
                          f"📄 File: {os.path.basename(actual_path)}\n"
                          f"📊 URLs exported: {total_urls:,}\n"
                          f"📏 File size: {file_size:,} bytes\n"
                          f"📁 Location: {actual_path}")
            
            messagebox.showinfo("Export Successful", success_msg)
            self.update_status(f"✅ Exported {total_urls} URLs to {os.path.basename(actual_path)}")
            
        except Exception as e:
            error_msg = f"Failed to export results: {str(e)}"
            messagebox.showerror("Export Error", error_msg)
            self.update_status(f"❌ Export failed: {str(e)}")

    def _export_csv(self, save_path):
        """Export results to CSV - simplified format"""
        with open(save_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow(["# URL Validation Results - Simplified"])
            writer.writerow(["# Generated by", f"{current_config.APP_NAME} v{current_config.APP_VERSION}"])
            writer.writerow(["# Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            writer.writerow(["# Source File", self.current_file_info.get('name', 'Unknown')])
            writer.writerow([])
            
            # Data header
            writer.writerow(["Status", "URL"])
            
            # Summary
            writer.writerow(["SUMMARY", f"Valid: {len(self.results['valid'])}, Invalid: {len(self.results['invalid'])}"])
            writer.writerow([])
            
            # Valid URLs
            for url in self.results['valid']:
                writer.writerow(["Valid", url])
            
            # Invalid URLs
            for url in self.results['invalid']:
                writer.writerow(["Invalid", url])

    def _export_text(self, save_path):
        """Export results to text format"""
        with open(save_path, 'w', encoding='utf-8') as f:
            f.write(f"URL VALIDATION RESULTS - SIMPLIFIED\n")
            f.write(f"{'='*50}\n\n")
            f.write(f"Generated by: {current_config.APP_NAME} v{current_config.APP_VERSION}\n")
            f.write(f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Source File: {self.current_file_info.get('name', 'Unknown')}\n\n")
            
            f.write(f"SUMMARY\n")
            f.write(f"{'-'*20}\n")
            f.write(f"Valid URLs: {len(self.results['valid']):,}\n")
            f.write(f"Invalid URLs: {len(self.results['invalid']):,}\n")
            f.write(f"Total URLs: {len(self.results['valid']) + len(self.results['invalid']):,}\n\n")
            
            if self.results['valid']:
                f.write(f"VALID URLS\n")
                f.write(f"{'-'*30}\n")
                for i, url in enumerate(self.results['valid'], 1):
                    f.write(f"{i:3d}. {url}\n")
                f.write(f"\n")
            
            if self.results['invalid']:
                f.write(f"INVALID URLS\n")
                f.write(f"{'-'*30}\n")
                for i, url in enumerate(self.results['invalid'], 1):
                    f.write(f"{i:3d}. {url}\n")
                f.write(f"\n")
            
            if not self.results['valid'] and not self.results['invalid']:
                f.write("No URLs found in the file.\n")

    def update_status(self, message):
        """Update status bar"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_bar.config(text=f"[{timestamp}] {message}")

    def on_closing(self):
        """Handle app closing"""
        if self.processing_thread and self.processing_thread.is_alive():
            if messagebox.askokcancel("Quit", "Processing is running. Quit anyway?"):
                self.validator.cancel_processing()
                self.root.destroy()
        else:
            self.root.destroy()


def main():
    """Main function"""
    try:
        root = tk.Tk()
        app = LinkValidatorApp(root)
        root.protocol("WM_DELETE_WINDOW", app.on_closing)
        root.mainloop()
        
    except Exception as e:
        import traceback
        print(f"Fatal error: {str(e)}")
        traceback.print_exc()
        try:
            messagebox.showerror("Fatal Error", f"Application failed to start: {str(e)}")
        except:
            pass


if __name__ == "__main__":
    main()
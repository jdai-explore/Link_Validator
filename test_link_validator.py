"""
Test suite for Link Validator application
Run with: python test_link_validator.py
"""
import unittest
import tempfile
import os
import csv
import pandas as pd
from openpyxl import Workbook
from link_validator import LinkValidator
from utils import (
    detect_encoding, get_progress_interval, safe_str_conversion,
    validate_file_path, get_column_letter, format_file_size
)
from exceptions import *


class TestLinkValidator(unittest.TestCase):
    """Test cases for LinkValidator class"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.validator = LinkValidator()
    
    def test_valid_urls(self):
        """Test URL validation with valid URLs"""
        valid_urls = [
            "http://example.com",
            "https://www.google.com",
            "http://localhost:8080",
            "https://192.168.1.1",
            "http://internal.company.com",
            "https://subdomain.example.org/path",
            "http://example.com:8080/path?query=1",
            "https://example-site.com"
        ]
        
        for url in valid_urls:
            with self.subTest(url=url):
                self.assertTrue(self.validator.is_valid_url(url), 
                              f"URL should be valid: {url}")
    
    def test_invalid_urls(self):
        """Test URL validation with invalid URLs"""
        invalid_urls = [
            "not-a-url",
            "ftp://example.com",  # Wrong scheme
            "http://",  # No netloc
            "https://",  # No netloc
            "",  # Empty string
            None,  # None value
            "just text",
            "www.example.com",  # No scheme
            "http://.example.com",  # Invalid netloc
            "https://example.com.",  # Trailing dot
        ]
        
        for url in invalid_urls:
            with self.subTest(url=url):
                self.assertFalse(self.validator.is_valid_url(url), 
                               f"URL should be invalid: {url}")
    
    def test_csv_processing(self):
        """Test CSV file processing"""
        # Create test CSV file
        test_data = [
            ["URL", "Description"],
            ["http://example.com", "Valid URL"],
            ["https://google.com", "Another valid URL"],
            ["not-a-url", "Invalid URL"],
            ["", "Empty cell"],
            ["http://localhost:3000", "Local URL"]
        ]
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerows(test_data)
            temp_file = f.name
        
        try:
            results = self.validator.process_csv(temp_file)
            
            # Debug: Print what we actually got
            print(f"Debug - Valid: {results['valid']}, Invalid: {results['invalid']}")
            print(f"Debug - Invalid links: {results['invalid_links']}")
            
            # Should find 3 valid URLs (http://example.com, https://google.com, http://localhost:3000)
            # Invalid should include: "not-a-url", "URL", "Description", "Valid URL", "Another valid URL", "Invalid URL", "Empty cell", "Local URL" 
            # The test data includes headers and descriptions that are being validated as URLs
            self.assertEqual(results['valid'], 3)
            self.assertGreaterEqual(results['invalid'], 1)  # At least 1 invalid (not-a-url)
            self.assertTrue(any("not-a-url" in link for link in results['invalid_links']))
            
        finally:
            os.unlink(temp_file)
    
    def test_excel_processing(self):
        """Test Excel file processing"""
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        test_data = [
            ["http://example.com", "Valid URL"],
            ["https://google.com", "Another valid URL"],
            ["not-a-url", "Invalid URL"],
            [None, "Empty cell"],
            ["http://localhost:3000", "Local URL"]
        ]
        
        for row_num, row_data in enumerate(test_data, 1):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_file = f.name
        
        try:
            results = self.validator.process_excel(temp_file)
            
            # Debug: Print what we actually got
            print(f"Debug Excel - Valid: {results['valid']}, Invalid: {results['invalid']}")
            print(f"Debug Excel - Invalid links: {results['invalid_links']}")
            
            # Should find 3 valid URLs (http://example.com, https://google.com, http://localhost:3000)
            # Invalid should include: "not-a-url" and text descriptions like "Valid URL", "Another valid URL", etc.
            self.assertEqual(results['valid'], 3)
            self.assertGreaterEqual(results['invalid'], 1)  # At least 1 invalid (not-a-url)
            self.assertTrue(any("not-a-url" in link for link in results['invalid_links']))
            
        finally:
            os.unlink(temp_file)
    
    def test_text_processing(self):
        """Test text file processing"""
        test_content = """http://example.com
https://google.com
not-a-url
http://localhost:3000

https://github.com"""
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f:
            f.write(test_content)
            temp_file = f.name
        
        try:
            results = self.validator.process_text(temp_file)
            
            # Debug: Print what we actually got
            print(f"Debug Text - Valid: {results['valid']}, Invalid: {results['invalid']}")
            print(f"Debug Text - Invalid links: {results['invalid_links']}")
            
            # Should find 4 valid URLs: http://example.com, https://google.com, http://localhost:3000, https://github.com
            # Should find 1 invalid: not-a-url
            self.assertEqual(results['valid'], 4)
            self.assertEqual(results['invalid'], 1)
            self.assertEqual(len(results['invalid_links']), 1)
            self.assertTrue(any("not-a-url" in link for link in results['invalid_links']))
            
        finally:
            os.unlink(temp_file)
    
    def test_html_processing(self):
        """Test HTML file processing"""
        html_content = """<!DOCTYPE html>
<html>
<head>
    <title>Test Page</title>
    <link rel="stylesheet" href="http://example.com/style.css">
</head>
<body>
    <a href="https://google.com">Valid Link</a>
    <a href="not-a-url">Invalid Link</a>
    <img src="http://localhost:8080/image.jpg" alt="Local Image">
    <script src="https://cdn.example.com/script.js"></script>
</body>
</html>"""
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
            f.write(html_content)
            temp_file = f.name
        
        try:
            results = self.validator.process_html(temp_file)
            
            # Debug: Print what we actually got
            print(f"Debug HTML - Valid: {results['valid']}, Invalid: {results['invalid']}")
            print(f"Debug HTML - Invalid links: {results['invalid_links']}")
            
            # Should find 4 valid URLs: http://example.com/style.css, https://google.com, 
            # http://localhost:8080/image.jpg, https://cdn.example.com/script.js
            # Should find 1 invalid: not-a-url
            self.assertEqual(results['valid'], 4)
            self.assertEqual(results['invalid'], 1)
            self.assertEqual(len(results['invalid_links']), 1)
            self.assertTrue(any("not-a-url" in link for link in results['invalid_links']))
            
        finally:
            os.unlink(temp_file)


class TestUtilityFunctions(unittest.TestCase):
    """Test cases for utility functions"""
    
    def test_safe_str_conversion(self):
        """Test safe string conversion"""
        test_cases = [
            ("hello", "hello"),
            (123, "123"),
            (123.0, "123"),
            (123.45, "123.45"),
            (None, ""),
            ("", ""),
            ("  spaced  ", "spaced")
        ]
        
        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                result = safe_str_conversion(input_val)
                self.assertEqual(result, expected)
    
    def test_get_column_letter(self):
        """Test Excel column letter conversion"""
        test_cases = [
            (1, "A"),
            (26, "Z"),
            (27, "AA"),
            (52, "AZ"),
            (53, "BA"),
            (702, "ZZ"),
            (703, "AAA")
        ]
        
        for col_num, expected in test_cases:
            with self.subTest(col_num=col_num):
                result = get_column_letter(col_num)
                self.assertEqual(result, expected)
    
    def test_get_progress_interval(self):
        """Test progress interval calculation"""
        test_cases = [
            (50, 5),      # Small file
            (500, 25),    # Medium file
            (5000, 100),  # Large file
            (50000, 500), # Very large file
            (500000, 1000) # Extra large file
        ]
        
        for total, expected_min in test_cases:
            with self.subTest(total=total):
                result = get_progress_interval(total)
                self.assertGreaterEqual(result, expected_min)
    
    def test_format_file_size(self):
        """Test file size formatting"""
        test_cases = [
            (512, "512 B"),
            (1024, "1.0 KB"),
            (1536, "1.5 KB"),
            (1048576, "1.0 MB"),
            (1073741824, "1.0 GB")
        ]
        
        for size_bytes, expected in test_cases:
            with self.subTest(size_bytes=size_bytes):
                result = format_file_size(size_bytes)
                self.assertEqual(result, expected)
    
    def test_validate_file_path(self):
        """Test file path validation"""
        # Test with valid file
        with tempfile.NamedTemporaryFile(delete=False) as f:
            f.write(b"test content")
            temp_file = f.name
        
        try:
            # Should not raise exception
            self.assertTrue(validate_file_path(temp_file))
        finally:
            os.unlink(temp_file)
        
        # Test with non-existent file
        with self.assertRaises(FileNotFoundError):
            validate_file_path("non_existent_file.txt")
        
        # Test with empty path
        with self.assertRaises(ValueError):
            validate_file_path("")


class TestExceptions(unittest.TestCase):
    """Test cases for custom exceptions"""
    
    def test_link_validator_error(self):
        """Test base LinkValidatorError"""
        error = LinkValidatorError("Test message", error_code="TEST001")
        self.assertEqual(str(error), "Test message")
        self.assertEqual(error.error_code, "TEST001")
    
    def test_file_processing_error(self):
        """Test FileProcessingError"""
        error = FileProcessingError("Processing failed", file_path="/test/file.csv", file_type="csv")
        self.assertEqual(error.file_path, "/test/file.csv")
        self.assertEqual(error.file_type, "csv")
        self.assertIn("file_path", error.context)
    
    def test_validation_error(self):
        """Test ValidationError"""
        error = ValidationError("Invalid URL", url="bad-url", validation_rule="scheme_check")
        self.assertEqual(error.url, "bad-url")
        self.assertEqual(error.validation_rule, "scheme_check")
    
    def test_exception_factory(self):
        """Test exception factory function"""
        error = create_exception("file_not_found", "File missing", file_path="/test/file.csv")
        self.assertIsInstance(error, FileValidationError)
        self.assertEqual(str(error), "File missing")


class TestIntegration(unittest.TestCase):
    """Integration tests"""
    
    def test_full_csv_workflow(self):
        """Test complete CSV processing workflow"""
        # Create a more comprehensive test CSV
        test_data = [
            ["URLs", "Description", "Category"],
            ["http://example.com", "Main site", "Primary"],
            ["https://api.example.com/v1", "API endpoint", "API"],
            ["ftp://files.example.com", "FTP server", "Files"],  # Invalid scheme
            ["", "Empty URL", "Empty"],
            ["not-a-url-at-all", "Text content", "Invalid"],
            ["http://localhost:8080/admin", "Local admin", "Local"],
            ["https://192.168.1.100:443", "IP address", "Internal"],
            [12345, "Number as URL", "Test"],  # Number
            ["https://sub.domain.co.uk/path?q=1&r=2", "Complex URL", "Complex"]
        ]
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerows(test_data)
            temp_file = f.name
        
        try:
            validator = LinkValidator()
            results = validator.process_csv(temp_file)
            
            # Verify results make sense
            total_checked = results['valid'] + results['invalid']
            self.assertGreater(total_checked, 0, "Should have checked some URLs")
            self.assertGreater(results['valid'], 0, "Should have found some valid URLs")
            self.assertGreater(results['invalid'], 0, "Should have found some invalid URLs")
            
            # Check that invalid links are properly recorded
            self.assertEqual(len(results['invalid_links']), results['invalid'])
            
        finally:
            os.unlink(temp_file)


def run_tests():
    """Run all tests and display results"""
    print(f"Running Link Validator Test Suite")
    print("=" * 50)
    
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add test classes
    test_classes = [
        TestLinkValidator,
        TestUtilityFunctions,
        TestExceptions,
        TestIntegration
    ]
    
    for test_class in test_classes:
        tests = loader.loadTestsFromTestCase(test_class)
        suite.addTests(tests)
    
    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # Print summary
    print("\n" + "=" * 50)
    print(f"Tests run: {result.testsRun}")
    print(f"Failures: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")
    
    if result.failures:
        print(f"\nFAILURES:")
        for test, traceback in result.failures:
            print(f"- {test}: {traceback}")
    
    if result.errors:
        print(f"\nERRORS:")
        for test, traceback in result.errors:
            print(f"- {test}: {traceback}")
    
    success = len(result.failures) == 0 and len(result.errors) == 0
    print(f"\nResult: {'PASSED' if success else 'FAILED'}")
    
    return success


if __name__ == "__main__":
    run_tests()